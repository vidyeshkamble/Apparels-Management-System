# import pandas as pd

# dataframe = pd.read_excel('Products3.xlsx')

# print(dataframe)

import openpyxl

# Load the Excel workbook
dataframe = openpyxl.load_workbook('Products3.xlsx')

# Read the active sheet
dataframe1 = dataframe.active

# Open the output text file in write mode
with open("file.txt", "w") as f:
    # Write the base SQL command
    table_name = "products"  # Change this to your actual table name
    f.write(f"INSERT INTO {table_name} (id, name, image_url, price) VALUES\n")
    
    # Iterate over rows and generate SQL insert values
    insert_values = []
    for row in range(2, dataframe1.max_row + 1):  # Assuming first row is a header
        row_data = []
        for col in range(1, dataframe1.max_column + 1):
            cell_value = dataframe1.cell(row=row, column=col).value
            if isinstance(cell_value, str):
                row_data.append(f"'{cell_value.replace('\'', '\'\'')}'")  # Escape single quotes for SQL
            elif cell_value is None:
                row_data.append("NULL")
            else:
                row_data.append(str(cell_value))
        insert_values.append(f"({', '.join(row_data)})")
    
    # Write all insert values separated by commas
    f.write(",\n".join(insert_values) + ";\n")
